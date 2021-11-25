import torch

from PIL import Image
import numpy as np
import openpyxl
from sklearn.decomposition import PCA

def dim_redu(data, gray, dim_ratio):
    """
    将三通道图片灰度处理后进行PCA降维
    cifar10图片原大小为32*32*3，展开为3072，总共包含3072个特征值，如果做灰度处理会减少为1024，然后按pca设定的比例进行降维
    Args:
        data:输入数据为四维[batch,H,W,C]
        gray:是否进行灰度处理
        dim_ratio:降维比率
    """
    cifar_train = data

    # 灰度处理
    if gray:
        print("灰度处理...")
        batch_size = len(cifar_train)
        cifar_gray = []
        for i in range(len(cifar_train)):
            gray_item = np.array(Image.fromarray(cifar_train[i]).convert('L'))
            cifar_gray.append(gray_item)
        cifar_train = np.array(cifar_gray).reshape(batch_size, -1)
        print("灰度处理完成")

    if dim_ratio == 1:
        cifar_train = cifar_train.reshape(len(cifar_train), -1)
    else:
        cifar_train = cifar_train.reshape(len(cifar_train), -1)
        # PCA降维
        print("开始PCA降维，降维比率为", dim_ratio, "...")
        pca_model = PCA(n_components=int(cifar_train.shape[-1] * dim_ratio))
        pca_model.fit(cifar_train)
        cifar_train = pca_model.transform(cifar_train)
        print("PCA降维完成")

    return cifar_train

def get_kfold_data(k, i, X, y):
    """
    K折验证
    Args:
        k:将原数据分为几份(原题目为10)
        i:当前所划分的验证数据为第几份
        X:数据集
        y:数据标签
    """
    fold_size = X.shape[0] // k  # 每份的个数:数据总条数/折数（组数）

    val_start = i * fold_size
    if i != k - 1:
        val_end = (i + 1) * fold_size
        X_valid, y_valid = X[val_start:val_end], y[val_start:val_end]
        X_train = torch.cat((X[0:val_start], X[val_end:]), dim=0)
        y_train = torch.cat((y[0:val_start], y[val_end:]), dim=0)
    else:  # 若是最后一折交叉验证
        X_valid, y_valid = X[val_start:], y[val_start:]  # 若不能整除，将多的case放在最后一折里
        X_train = X[0:val_start]
        y_train = y[0:val_start]

    return X_train, y_train, X_valid, y_valid

def WriteExcelRow(data, path, colIndex, sheet):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    if colIndex == -1:
        colIndex = ws.max_row + 1
    for i in range(len(data)):
        ws.cell(row=colIndex, column=i+1).value = data[i]
    wb.save(path)
